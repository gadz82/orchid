"""
Pluggable document parsers — extract text from various file formats.

Each parser implements ``DocumentParser.parse()`` which takes raw bytes
and returns extracted text. New formats can be added by subclassing
and registering in ``PARSER_REGISTRY``.

The PDF / DOCX / XLSX / image parsers depend on heavy third-party
packages (``pymupdf``, ``python-docx``, ``openpyxl``, ``Pillow``) that
ship via the ``[documents]`` extra:

    pip install orchid-ai[documents]

The base :class:`DocumentParser` plus the lightweight
:class:`CSVParser` and :class:`TextParser` are always available and
require no extras.  Each heavy parser raises a friendly
:class:`ImportError` from inside ``parse()`` if its underlying
package is missing — so importing this module never fails on a lean
``pip install orchid-ai`` install.
"""

from __future__ import annotations

import csv
import io
import logging
from abc import ABC, abstractmethod
from pathlib import Path

logger = logging.getLogger(__name__)


_DOCUMENTS_EXTRA_HINT = (
    "Install via `pip install orchid-ai[documents]` to enable "
    "PDF / DOCX / XLSX parsing, or install the underlying package "
    "directly."
)


def _missing_extra(parser_name: str, package: str, distribution: str) -> ImportError:
    """Build a uniform, actionable :class:`ImportError`.

    ``parser_name`` is the human-readable parser label
    (``"PDF"``, ``"DOCX"``, ``"XLSX"``).  ``package`` is the import
    name (``"fitz"``, ``"docx"``, …).  ``distribution`` is the PyPI
    distribution that owns the import (``"pymupdf"``, ``"python-docx"``,
    …) — the two often differ and we want the message to spell out
    exactly what to ``pip install``.
    """
    return ImportError(
        f"{parser_name} parsing requires the '{package}' package "
        f"(PyPI: '{distribution}'), which is not installed.\n"
        f"{_DOCUMENTS_EXTRA_HINT}"
    )


class DocumentParser(ABC):
    """Base class for document parsers."""

    @abstractmethod
    async def parse(self, file_bytes: bytes, filename: str) -> str:
        """Extract text from file bytes. Returns the extracted text."""
        ...


class PDFParser(DocumentParser):
    """Parse PDF files using PyMuPDF (fitz).

    Requires the ``pymupdf`` package (ships with
    ``orchid-ai[documents]``).  Raises a friendly
    :class:`ImportError` from :meth:`parse` if the package is missing
    so importing this module on a lean install never fails.
    """

    async def parse(self, file_bytes: bytes, filename: str) -> str:
        try:
            import fitz  # pymupdf
        except ImportError as exc:  # pragma: no cover — exercised only when dep missing
            raise _missing_extra("PDF", "fitz", "pymupdf") from exc

        doc = fitz.open(stream=file_bytes, filetype="pdf")
        pages: list[str] = []
        for page in doc:
            pages.append(page.get_text())
        doc.close()
        return "\n\n".join(pages)


class DOCXParser(DocumentParser):
    """Parse Word documents using python-docx.

    Requires the ``python-docx`` package (ships with
    ``orchid-ai[documents]``).  Raises a friendly
    :class:`ImportError` from :meth:`parse` if the package is missing.
    """

    async def parse(self, file_bytes: bytes, filename: str) -> str:
        try:
            from docx import Document
        except ImportError as exc:  # pragma: no cover — exercised only when dep missing
            raise _missing_extra("DOCX", "docx", "python-docx") from exc

        doc = Document(io.BytesIO(file_bytes))
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        return "\n\n".join(paragraphs)


class XLSXParser(DocumentParser):
    """Parse Excel spreadsheets using openpyxl, outputting as CSV-like text.

    Requires the ``openpyxl`` package (ships with
    ``orchid-ai[documents]``).  Raises a friendly
    :class:`ImportError` from :meth:`parse` if the package is missing.
    """

    async def parse(self, file_bytes: bytes, filename: str) -> str:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover — exercised only when dep missing
            raise _missing_extra("XLSX", "openpyxl", "openpyxl") from exc

        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        sheets: list[str] = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows: list[str] = []
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                if any(cells):
                    rows.append(" | ".join(cells))
            if rows:
                sheets.append(f"## Sheet: {sheet_name}\n" + "\n".join(rows))

        wb.close()
        return "\n\n".join(sheets)


class CSVParser(DocumentParser):
    """Parse CSV files."""

    async def parse(self, file_bytes: bytes, filename: str) -> str:
        text = file_bytes.decode("utf-8", errors="replace")
        reader = csv.reader(io.StringIO(text))
        rows: list[str] = []
        for row in reader:
            rows.append(" | ".join(row))
        return "\n".join(rows)


class TextParser(DocumentParser):
    """Passthrough parser for plain text and markdown files."""

    async def parse(self, file_bytes: bytes, filename: str) -> str:
        return file_bytes.decode("utf-8", errors="replace")


class ImageParser(DocumentParser):
    """
    Extract text from images using LiteLLM vision model.

    The model is configurable via ``vision_model`` (constructor param or
    ``VISION_MODEL`` env var). Falls back to a placeholder if extraction fails.
    """

    def __init__(self, vision_model: str = ""):
        self._vision_model = vision_model

    async def parse(self, file_bytes: bytes, filename: str) -> str:
        import base64

        import litellm

        model = self._vision_model
        if not model:
            return f"[Image file: {filename} — no vision model configured]"

        ext = Path(filename).suffix.lower().lstrip(".")
        mime = f"image/{ext}" if ext in ("png", "jpg", "jpeg", "gif", "webp") else "image/png"
        b64 = base64.b64encode(file_bytes).decode("ascii")

        logger.info(
            "[ImageParser] Parsing %s (%d bytes, mime=%s, b64_len=%d) with model=%s",
            filename,
            len(file_bytes),
            mime,
            len(b64),
            model,
        )

        try:
            from ..llm import get_llm_kwargs

            response = await litellm.acompletion(
                model=model,
                messages=[
                    {
                        "role": "user",
                        "content": [
                            {
                                "type": "text",
                                "text": "Extract all text and describe the content of this image in detail.",
                            },
                            {"type": "image_url", "image_url": {"url": f"data:{mime};base64,{b64}"}},
                        ],
                    }
                ],
                temperature=0,
                **get_llm_kwargs(model),
            )
            result = response.choices[0].message.content or f"[Image: {filename}]"
            logger.info(
                "[ImageParser] Result for %s (first 300 chars): %s",
                filename,
                result[:300],
            )
            return result
        except Exception as exc:
            logger.warning("[ImageParser] Vision extraction failed for %s: %s", filename, exc)
            return f"[Image file: {filename} — vision extraction failed]"


# ── Parser registry ──────────────────────────────────────────

PARSER_REGISTRY: dict[str, type[DocumentParser]] = {
    ".pdf": PDFParser,
    ".docx": DOCXParser,
    ".xlsx": XLSXParser,
    ".csv": CSVParser,
    ".md": TextParser,
    ".txt": TextParser,
    ".png": ImageParser,
    ".jpg": ImageParser,
    ".jpeg": ImageParser,
}

SUPPORTED_EXTENSIONS = set(PARSER_REGISTRY.keys())


def register_parser(ext: str, cls: type[DocumentParser]) -> None:
    """
    Register a custom parser for a file extension (OCP — extend without editing).

    Parameters
    ----------
    ext : str
        File extension including the dot (e.g. ``".pptx"``).
    cls : type[DocumentParser]
        Parser class to handle files with this extension.
    """
    PARSER_REGISTRY[ext] = cls
    SUPPORTED_EXTENSIONS.add(ext)
    logger.info("[Parsers] Registered %s for extension '%s'", cls.__name__, ext)


def get_parser(filename: str, *, vision_model: str = "") -> DocumentParser:
    """Get the appropriate parser for a file by extension."""
    ext = Path(filename).suffix.lower()
    cls = PARSER_REGISTRY.get(ext)
    if not cls:
        raise ValueError(f"Unsupported file type: {ext}. Supported: {', '.join(sorted(SUPPORTED_EXTENSIONS))}")
    if cls is ImageParser:
        return cls(vision_model=vision_model)
    return cls()
